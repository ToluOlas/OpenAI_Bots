import os
import openai
from dotenv import load_dotenv
from colorama import Fore, Back, Style
import pandas as pd
from openpyxl import load_workbook

#find excel file path
file_path = 'C:\\Users\\4939921\\OneDrive - MyFedEx\\Documents\\command_line_chatgpt-main\\OAR_test_dataset_new.xlsx'
# Load the existing Excel file
workbook = load_workbook(file_path)
sheet = workbook.active
row = 2

# 2D array - x
# 0 - Package Tracking, 1 - Customs & Clearance, 2 - Billing, 3 - Pickup, 
# 4 - Rating & Transit Times, 5 - Account & Sales Support, 6 - POD proof of delivery, 
# 7 - Claims, Supplies, 8 - IT Support, 9 - Other)

# 2D array - y 
# 0 = correctly guessed / 1 = incorrectly guessed
category_accuracy = [
    [0, 0],  # category_accuracy[0] corresponds to "package tracking"
    [0, 0],  # category_accuracy[1] corresponds to "customs & clearance"
    [0, 0],  # category_accuracy[2] corresponds to "billing"
    [0, 0],  # category_accuracy[3] corresponds to "pickup"
    [0, 0],  # category_accuracy[4] corresponds to "rating & transit times"
    [0, 0],  # category_accuracy[5] corresponds to "account & sales support"
    [0, 0],  # category_accuracy[6] corresponds to "pod proof of delivery"
    [0, 0],  # category_accuracy[7] corresponds to "claims"
    [0, 0],  # category_accuracy[8] corresponds to "supplies"
    [0, 0],  # category_accuracy[9] corresponds to "it support"
    [0, 0]   # category_accuracy[10] corresponds to "other"
]

#evaluation
while sheet.cell(row, 1).value is not None:

    category_correct = (sheet.cell(row, 2).value).lower() #column 2 = Primary Categories
    category_guess = (sheet.cell(row, 4).value).lower()
    print ("Guessed: "+category_guess)
    print ("Answer: "+category_correct)

    if category_correct == category_guess:
        is_correct = 0 # since 0 = correct
        print("Correct")
    else:
        is_correct = 1 # since 1 = incorrect
        print("Incorrect")

    if category_correct == 'package tracking':
        category_accuracy[0][is_correct] += 1
    elif category_correct == "customs & clearance":
        category_accuracy[1][is_correct] += 1
    elif category_correct == "billing":
        category_accuracy[2][is_correct] += 1
    elif category_correct == "pickup":
        category_accuracy[3][is_correct] += 1
    elif category_correct == "rating & transit times":
        category_accuracy[4][is_correct] += 1
    elif category_correct == "account & sales support":
        category_accuracy[5][is_correct] += 1
    elif category_correct == "pod proof of delivery":
        category_accuracy[6][is_correct] += 1
    elif category_correct == "claims":
        category_accuracy[7][is_correct] += 1
    elif category_correct == "supplies":
        category_accuracy[8][is_correct] += 1
    elif category_correct == "it support":
        category_accuracy[9][is_correct] += 1
    elif category_correct == "other":
        category_accuracy[10][is_correct] += 1
    else:
        print("Category not found in mapping.")

    row += 1

#save results

#save category guess
sheet.cell(2, 7).value = category_accuracy[0][0]
sheet.cell(2, 8).value = category_accuracy[0][1]

sheet.cell(3, 7).value = category_accuracy[1][0]
sheet.cell(3, 8).value = category_accuracy[1][1]

sheet.cell(4, 7).value = category_accuracy[2][0]
sheet.cell(4, 8).value = category_accuracy[2][1]

sheet.cell(5, 7).value = category_accuracy[3][0]
sheet.cell(5, 8).value = category_accuracy[3][1]

sheet.cell(6, 7).value = category_accuracy[4][0]
sheet.cell(6, 8).value = category_accuracy[4][1]

sheet.cell(7, 7).value = category_accuracy[5][0]
sheet.cell(7, 8).value = category_accuracy[5][1]

sheet.cell(8, 7).value = category_accuracy[6][0]
sheet.cell(8, 8).value = category_accuracy[6][1]

sheet.cell(9, 7).value = category_accuracy[7][0]
sheet.cell(9, 8).value = category_accuracy[7][1]

sheet.cell(10, 7).value = category_accuracy[8][0]
sheet.cell(10, 8).value = category_accuracy[8][1]

sheet.cell(11, 7).value = category_accuracy[9][0]
sheet.cell(11, 8).value = category_accuracy[9][1]

sheet.cell(12, 7).value = category_accuracy[10][0]
sheet.cell(12, 8).value = category_accuracy[10][1]

# Save the workbook 
workbook.save(file_path)






