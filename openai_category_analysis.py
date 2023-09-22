import os
import openai
from dotenv import load_dotenv
from colorama import Fore, Back, Style
import pandas as pd
from openpyxl import load_workbook

#find excel file path
file_path = 'C:\\Users\\4939921\\OneDrive - MyFedEx\\Documents\\command_line_chatgpt-main\\OAR_test_dataset_new.xlsx'
# Load the existing Excel file
workbook = load_workbook(file_path)
sheet = workbook.active

# load values from the .env file if it exists
load_dotenv()


# configure OpenAI
openai.api_key = os.getenv("OPENAI_API_KEY")

PRIMARY_INSTRUCTIONS = """You are a bot that reads emails, and return the most appropriate catergory.
You will be given an email, and must choose a category that is the most fitting for the text.
An email MUST have one category.

The possible categories are: Package Tracking, Customs & Clearance, Billing, Pickup, Rating & Transit Times, Account & Sales Support, POD proof of delivery, Claims, Supplies, IT Support, Other

Do NOT use any other categories other than the ones that are listed.
Ouput in this format: [INSERT CATEGORY]
Do not add any other text that isn't the chosen category.

EXAMPLES:

Input: The parcel 784573749880 has been stuck in delivery with the following message since 15th March 2020:  "Clearance delay – Import Clearance instructions from the importer are required."  Please can someone let us know what is required so we can get this delivered as soon as possible?
Bot: Package Tracking

Input: Hello,  I hope you're well. I'm trying to reach out on the phone but no one is picking up. I am urgently waiting to receive my package, it's been stuck in Paris since Tuesday and I'm wondering what's going on? If you could advise us on this as soon as possible, that would be great!  Thank you.
Bot: Custom & Clearance

Input: We have been charged an extra 9 Euro when the measurements do not exceed what is this for?
Bot: Billing

Input: I need quotes to send some boxes to very national desintations.
Bot: Rating & Transit Times

Input: I was charged more that my shipping cost and didnt get my receipt.
Bot: Other

Input: Please confirm that the shipment will be collected today from our warehouse.
Bot: Pickup

Input: I would like to delete my account with FedEx so as to stop you making unauthorised charges. Please tell me how to delete my FedEx account.
Bot: Account & Sales Support

Input: Please send me the hard copy proof of delivery for the below shipments.
Bot: POD Proof of Delivery

Input: This says my Parcel was delivered, but I don't have it. It says it was left outside the front door on March 15th at 12:35 we both work from home full time and were home at that time. Who have you left my package with?
Bot: Claims

Input: We need to request some more stationery please can you send us.
Bot: Supplies

Input: I received a response form the Other email address which suggested I try phoning the IT Helpdesk on 01827 711611 
Bot: IT Support

Input: Good afternoon, Could you please help me with my request could you please forward me a copy of the C88 SAD for this consignment, Collection address: Schaltbau Machine Electrics Ltd, 335-336 Woodside Way, Springvale Industrial Estate, Woodside Way, Cwmbran. NP44 5BR. UK. Delivery address: Asseco CEIT, a. s. Rosina 928 Rosina 013 22 SLOVAKIA Thank you
Bot: Custom & Clearance

Input: Hi, Could you provide export docs for consignment 326587198, as soon as they are available, please? Booked on account 002424788 Thanks
Bot: Custom & Clearance

Input: ensure that this package will be picked up tomorrow it was due to be picked up from us today driver never arrived
Bot: Pickup

Input: Hi I have received and SMS asking me to pay a charge whoever it is for a parcel i never received. I have even got  a refund back from the company in question. Please advise why i am still being charged this. Thanks
Bot: Claims"""

TEMPERATURE = 0.5
MAX_TOKENS = 200
FREQUENCY_PENALTY = 0.2
PRESENCE_PENALTY = 0.6

def get_response(instructions, new_question):
    """Get a response from ChatCompletion

    Args:
        instructions: The instructions for the chat bot - this determines how it will behave
        new_question: The new question to ask the bot

    Returns:
        The response text
    """
    # build the messages
    messages = [
        { "role": "system", "content": instructions },
    ]
    # add the new question
    messages.append({ "role": "user", "content": new_question })

    completion = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=messages,
        temperature=TEMPERATURE,
        max_tokens=MAX_TOKENS,
        top_p=1,
        frequency_penalty=FREQUENCY_PENALTY,
        presence_penalty=PRESENCE_PENALTY,
    )
    return completion.choices[0].message.content

def main():
    os.system("cls" if os.name == "nt" else "clear")
    # Introduction text
    print(Fore.CYAN + Style.BRIGHT + "Bot is now running...")

    #while loop ends when row without a guess from OpenAI is found
    row = 1
    while sheet.cell(row, 4).value is not None:
        row += 1

    while sheet.cell(row, 1).value is not None:

        # retrieve email message 
        new_question = "'" + sheet.cell(row, 1).value + "'"
        print (new_question)

        #retrieve OpenAI's guess for the category and print it
        primary_category = get_response(PRIMARY_INSTRUCTIONS, new_question)
        print(Fore.CYAN + Style.BRIGHT + "OpenAI has guessed: " + Style.NORMAL + primary_category)

        #save category guess
        sheet.cell(row, 4).value = primary_category

        #increment row
        row = row + 1

        # Save at multiples of 10
        if row % 10 == 0:
            workbook.save(file_path)

    #save file
    workbook.save(file_path)

if __name__ == "__main__":
    main()
