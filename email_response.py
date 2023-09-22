import os
import openai
from dotenv import load_dotenv
from colorama import Fore, Back, Style
import pandas as pd
from openpyxl import load_workbook

file_path = 'C:\\Users\\4939921\\OneDrive - MyFedEx\\Documents\\command_line_chatgpt-main\\OAR_test_dataset_new.xlsx'

# Load the existing Excel file
workbook = load_workbook(file_path)
sheet = workbook.active

# Find the first available cell in the first column
row = 1
while sheet.cell(row, 1).value is not None:
    row += 1

# load values from the .env file if it exists
load_dotenv()

# configure OpenAI
openai.api_key = os.getenv("OPENAI_API_KEY")

INSTRUCTIONS = """You are a bot that generates messages for a dataset that will be used to train a machine learning program. 
The emails you generate will be customer queries, related to Billing.
Pretend to be a customer sending a message to FedEx, with an issue or request related to billing.
Limit the message to one paragraph, no line breaks.
Do not list a subject, greeting, signature or anything else other than the contents of the message.
Do not send anything else other than the generated message.
The message does not need to be formal, tone may be upset, confused, annoyed, angry etc.
Try to avoid the explicit use of the word "Billing".

Examples (do not use information like exact account numbers and values from these examples when generating messages):

"Consignment number 317573616 was contracted with a global service and to this day we have not received it yet when you issue the invoice we will not pay for the global rate at best we would understand if it was invoiced as economy "

"I received the below email on the 30th July however I still have not received the credit in the account"

"Hi, Could you send me a statement of account and a copy of all open invoices for this account 002638851 as a matter of urgency. We have been billed to this account in the past, but we should not be. All shipments are booked under account 54451 and that is the account we should be billed under. But if you send me these invoices I will deal with them as a matter of urgency and then you can close account 002638851 Many thanks"

"the correct weight of said shipment both by your means and by ours is 31Kg and not the invoiced of 67Kg I beg you to proceed to confirm a change in the amount of said invoice "

"We have observed that we have a direct debit receipt from you for an amount of € 108 56 and we do not know what it corresponds to"

"Why is there a difference between our customs payment and the invoice we receive? I enclose the invoice and proof of payment"

"Hi All, Please can someone advise on why this has been invoiced to the shipper and not the receiver? Thank you"

"Hi, Could you send me a statement of account and a copy of all open invoices for this account 002638851 as a matter of urgency. We have been billed to this account in the past, but we should not be. All shipments are booked under account 54451 and that is the account we should be billed under. But if you send me these invoices I will deal with them as a matter of urgency and then you can close account 002638851 Many thanks" """

TEMPERATURE = 0.5
MAX_TOKENS = 500
FREQUENCY_PENALTY = 0.2
PRESENCE_PENALTY = 0.6
# limits how many questions we include in the prompt
MAX_CONTEXT_QUESTIONS = 10


def get_response(instructions, new_question):
    """Get a response from ChatCompletion

    Args:
        instructions: The instructions for the chat bot - this determines how it will behave
        previous_questions_and_answers: Chat history
        new_question: The new question to ask the bot

    Returns:
        The response text
    """
    # build the messages
    messages = [
        { "role": "system", "content": instructions },
    ]
    # add the new question
    messages.append({ "role": "user", "content": new_question })

    completion = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=messages,
        temperature=TEMPERATURE,
        max_tokens=MAX_TOKENS,
        top_p=1,
        frequency_penalty=FREQUENCY_PENALTY,
        presence_penalty=PRESENCE_PENALTY,
    )
    return completion.choices[0].message.content


def get_moderation(question):
    """
    Check the question is safe to ask the model

    Parameters:
        question (str): The question to check

    Returns a list of errors if the question is not safe, otherwise returns None
    """

    errors = {
        "hate": "Content that expresses, incites, or promotes hate based on race, gender, ethnicity, religion, nationality, sexual orientation, disability status, or caste.",
        "hate/threatening": "Hateful content that also includes violence or serious harm towards the targeted group.",
        "self-harm": "Content that promotes, encourages, or depicts acts of self-harm, such as suicide, cutting, and eating disorders.",
        "sexual": "Content meant to arouse sexual excitement, such as the description of sexual activity, or that promotes sexual services (excluding sex education and wellness).",
        "sexual/minors": "Sexual content that includes an individual who is under 18 years old.",
        "violence": "Content that promotes or glorifies violence or celebrates the suffering or humiliation of others.",
        "violence/graphic": "Violent content that depicts death, violence, or serious physical injury in extreme graphic detail.",
    }
    response = openai.Moderation.create(input=question)
    if response.results[0].flagged:
        # get the categories that are flagged and generate a message
        result = [
            error
            for category, error in errors.items()
            if response.results[0].categories[category]
        ]
        return result
    return None


def main():
    os.system("cls" if os.name == "nt" else "clear")
    # Introduction text
    print(Fore.CYAN + Style.BRIGHT + "Bot is now running...")
    while True:
        # ask the user for their question
        new_question = input(
            Fore.GREEN + Style.BRIGHT + "Enter email here: " + Style.RESET_ALL
        )
        # check the question is safe
        errors = get_moderation(new_question)
        if errors:
            print(
                Fore.RED
                + Style.BRIGHT
                + "Sorry, you're question didn't pass the moderation check:"
            )
            for error in errors:
                print(error)
            print(Style.RESET_ALL)
            continue
        response = get_response(INSTRUCTIONS, new_question)

        # print the response
        print(Fore.CYAN + Style.BRIGHT + "BOT: " + Style.NORMAL + response)

        # Write the text into the first available cell in the first column
        sheet.cell(row, 1).value = response
        sheet.cell(row, 2).value = "Billing"
        workbook.save(file_path)


if __name__ == "__main__":
    main()
